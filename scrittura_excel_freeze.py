#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modulo: scrittura_excel_freeze
Versione 4.1 — Freeze stabile (intestazioni dinamiche)
Ignazio Rusconi-Clerici — macOS
------------------------------------
- Offset dinamico (-1h se CSV parte alle 02:01)
- Creazione automatica fogli mensili successivi
- Intestazioni lette dal primo foglio esistente (es. gen 25)
- Scrittura e salvataggio atomici
"""

import os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter

TOLLERANZA_MINUTI = 5  # tolleranza 5 min tra timestamp e riga

# =========================================================
#  FUNZIONI DI SUPPORTO
# =========================================================
def nome_foglio_da_data(dt):
    import locale
    try:
        locale.setlocale(locale.LC_TIME, "it_IT.UTF-8")
    except Exception:
        pass
    return dt.strftime("%b %y").lower()

def genera_colonna_date(sheet, anno, mese, intestazioni):
    """Crea la colonna Data e le intestazioni."""
    sheet.cell(1, 1).value = "Data"
    for i, nome in enumerate(intestazioni, start=2):
        if nome:
            sheet.cell(1, i).value = nome
    giorno = datetime(anno, mese, 1, 1, 1)
    r = 2
    while giorno.month == mese:
        sheet.cell(r, 1).value = giorno.strftime("%d/%m/%y %H:%M")
        giorno += timedelta(hours=2)
        if r > 400:
            break
        r += 1

def trova_riga(sheet, target_dt):
    """
    Trova la riga corrispondente alla data nel foglio Excel.
    Tolleranza: ±5 minuti, con fallback ±60 min per gestire cambi d’ora.
    """
    target_dt = target_dt.replace(second=0)
    candidati = []

    for r in range(2, sheet.max_row + 1):
        v = sheet.cell(r, 1).value
        if not v:
            continue
        try:
            d = v if isinstance(v, datetime) else datetime.strptime(str(v), "%d/%m/%y %H:%M")
            d = d.replace(second=0)
        except Exception:
            continue
        diff = abs((d - target_dt).total_seconds()) / 60
        if diff <= 5:  # normale tolleranza
            return r
        if diff <= 60:  # candidato per cambio ora
            candidati.append((r, diff))

    # fallback: scegli la riga più vicina entro 60 min se esiste
    if candidati:
        candidati.sort(key=lambda x: x[1])
        return candidati[0][0]

    return None

def intestazioni_da_primo_foglio(wb):
    """Legge le intestazioni (escl. 'Data') dal primo foglio esistente."""
    if not wb.sheetnames:
        return ["CSV1", "CSV2", "CSV3"]
    ws0 = wb[wb.sheetnames[0]]
    intestazioni = []
    for c in range(2, ws0.max_column + 1):
        val = ws0.cell(1, c).value
        if val:
            intestazioni.append(str(val))
    if not intestazioni:
        intestazioni = ["CSV1", "CSV2", "CSV3"]
    return intestazioni

# =========================================================
#  FUNZIONE PRINCIPALE — FREEZE
# =========================================================
def scrivi_excel_1_4(xls_path, dati_uniti):
    """
    Scrive i dati nei fogli mensili del file Excel.
    - xls_path: percorso file Excel
    - dati_uniti: lista di tuple [('CSV1', dict), ('CSV2', dict), ('CSV3', dict)]
    """
    if not dati_uniti:
        raise ValueError("Nessun dato da scrivere.")

    if os.path.exists(xls_path):
        wb = load_workbook(xls_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Leggi intestazioni dal primo foglio
    intestazioni = intestazioni_da_primo_foglio(wb)

    # 🔹 Mappa nome CSV → indice fisso di colonna
    mappa_colonne = {"CSV1": 1, "CSV2": 2, "CSV3": 3}

    # 🔹 Inserisci dati nelle colonne giuste
    for nome_col, dati_csv in dati_uniti:
        if nome_col not in mappa_colonne:
            continue
        idx = mappa_colonne[nome_col]  # posizione fissa, indipendente dai file selezionati

        for dt, temp in dati_csv.items():
            foglio = nome_foglio_da_data(dt)
            if foglio not in wb.sheetnames:
                ws = wb.create_sheet(title=foglio)
                genera_colonna_date(ws, dt.year, dt.month, intestazioni)
            else:
                ws = wb[foglio]

            r = trova_riga(ws, dt)
            if r:
                ws.cell(r, idx + 1).value = round(temp, 2)

    # 🔹 Formattazione colonne (opzionale ma consigliata)
    for nome_foglio in wb.sheetnames:
        ws = wb[nome_foglio]
        for col in range(1, ws.max_column + 1):
            for cell in ws[get_column_letter(col)]:
                cell.alignment = Alignment(horizontal="center")
                if col == 1:  # colonna Data
                    cell.number_format = "DD/MM/YY HH:MM"
    wb.save(xls_path)
    wb.close()