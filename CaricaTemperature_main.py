#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CaricaTemperature_main.py — Freeze stabile 1.4.2 (+ patch base_path + reset CSV)

Ignazio Rusconi-Clerici — macOS

Usa i moduli congelati:
- gui_freeze (crea_gui_1_4_2)
- lettura_csv_freeze (leggi_csv_4_0)
- merge_dati_freeze (presente nel freeze; non sempre necessario qui)
- scrittura_excel_freeze (scrivi_excel_1_4)

Patch richieste:
1) --base_path (default: ~/Dropbox/Documenti) usato come cartella iniziale per scegliere Excel/CSV,
   e salvato in un JSON di configurazione in ~/Library/CloudStorage/Dropbox/Documenti_IRC/Python/_Config.
2) Quando si seleziona un NUOVO Excel, vengono azzerati i path completi dei CSV selezionati
   nella sessione precedente (state reset), mantenendo solo i nomi attesi letti da B1–D1.
"""

from __future__ import annotations

# --- IRC shared bootstrap ---
# Rende disponibili i moduli in Python/shared/ senza dipendere da PYTHONPATH.
# Saltato se eseguito da bundle PyInstaller (sys.frozen=True): in quel caso
# i moduli sono gia' inclusi nel bundle.
import sys as _sys
from pathlib import Path as _Path
if not getattr(_sys, 'frozen', False):
    _shared = _Path.home() / "Library/CloudStorage/Dropbox/Documenti_IRC/Python/shared"
    if str(_shared) not in _sys.path:
        _sys.path.insert(0, str(_shared))
# --- end IRC shared bootstrap ---


APP_NAME = "CaricaTemperature"
VERSION  = "1.5.0"

from irc_paths import DROPBOX_ROOT, DROPBOX_USER_ROOT, app_config_dir
from irc_logging import setup_app_logger

import argparse
import json
import os
import threading
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Callable, Dict, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook

import gui_freeze
from path_widgets import log_path
import lettura_csv_freeze
import merge_dati_freeze  # keep import (freeze)
import scrittura_excel_freeze

log = setup_app_logger(APP_NAME, also_to_console=False)

CONFIG_DIR  = app_config_dir(APP_NAME)
CONFIG_FILE = CONFIG_DIR / "config.json"

# default richiesto: cartella "Documenti" sotto Dropbox
DEFAULT_BASE_PATH = str(DROPBOX_USER_ROOT / "Documenti")

# default richiesti per i chooser (Excel e CSV)
DEFAULT_EXCEL_DIR = DROPBOX_ROOT / "Case/Controlli/Temperature"
DEFAULT_CSV_DIR   = Path.home() / "Downloads"

# ============================================================
#  CONFIG
# ============================================================

@dataclass
class Settings:
    base_path: str = DEFAULT_BASE_PATH
    last_excel_dir: str = DEFAULT_EXCEL_DIR
    last_csv_dir: str = DEFAULT_CSV_DIR
def _ensure_dirs() -> None:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)


def load_settings() -> Settings:
    _ensure_dirs()
    s = Settings()
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                if data.get("base_path"):
                    s.base_path = str(data["base_path"])
                if data.get("last_excel_dir"):
                    s.last_excel_dir = str(data["last_excel_dir"])
                if data.get("last_csv_dir"):
                    s.last_csv_dir = str(data["last_csv_dir"])
        except Exception:
            # config corrotta: ignora e usa default
            pass
    return s


def save_settings(s: Settings) -> None:
    _ensure_dirs()
    CONFIG_FILE.write_text(json.dumps(asdict(s), indent=2, ensure_ascii=False), encoding="utf-8")


# ============================================================
#  HELPERS
# ============================================================

def read_expected_csv_names_from_excel(xls_path: Path) -> Dict[str, str]:
    """
    Legge i nomi attesi dei CSV dall'intestazione del primo foglio:
      B1 -> CSV1
      C1 -> CSV2
      D1 -> CSV3
    Se una cella è vuota, ritorna un placeholder.
    """
    wb = load_workbook(filename=str(xls_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = {
        "CSV1": ws["B1"].value,
        "CSV2": ws["C1"].value,
        "CSV3": ws["D1"].value,
    }
    out: Dict[str, str] = {}
    for k, v in raw.items():
        if v is None:
            out[k] = f"{k} (nome non trovato in intestazione)"
        else:
            s = str(v).strip()
            out[k] = s if s else f"{k} (nome non trovato in intestazione)"
    return out


def safe_basename(p: str) -> str:
    try:
        return Path(p).name
    except Exception:
        return p


# ============================================================
#  MAIN APP
# ============================================================

def main() -> None:
    args = parse_args()
    settings = load_settings()
    log.info(f"Avviato v{VERSION} — base_path: {settings.base_path}")

    # override base_path da CLI (se fornito)
    if args.base_path:
        settings.base_path = args.base_path

    # --- stato runtime ---
    stato = {
        "excel_path": None,  # type: Optional[str]
        "csv_paths": {"CSV1": None, "CSV2": None, "CSV3": None},  # type: Dict[str, Optional[str]]
        "csv_expected": {"CSV1": "CSV 1", "CSV2": "CSV 2", "CSV3": "CSV 3"},  # type: Dict[str, str]
        "log_file": None,  # type: Optional[Path]
    }

    app = tk.Tk()
    app.title("Caricamento Temperature — Freeze 1.4.2")
    app.minsize(820, 520)

    # -----------------------
    # UI-safe callbacks
    # -----------------------
    def ui_log(msg: str) -> None:
        def _write():
            try:
                app.txt_log.insert("end", msg.rstrip() + "\n")
                app.txt_log.see("end")
            except Exception:
                pass

            # log su file (accanto all'Excel), se attivo e se disponibile
            if getattr(app, "log_active", None) is not None and app.log_active.get():
                lf = stato.get("log_file")
                if lf:
                    try:
                        lf.parent.mkdir(parents=True, exist_ok=True)
                        with lf.open("a", encoding="utf-8") as f:
                            f.write(msg.rstrip() + "\n")
                    except Exception:
                        pass

        app.after(0, _write)

    def ui_status(msg: str) -> None:
        def _w():
            try:
                app.status_var.set(msg)
            except Exception:
                pass
        app.after(0, _w)

    def ui_progress(pct: int) -> None:
        pct = max(0, min(100, int(pct)))
        def _w():
            try:
                app.pbar["value"] = pct
            except Exception:
                pass
        app.after(0, _w)

    # -----------------------
    # Core actions
    # -----------------------
    def reset_csv_state_keep_expected() -> None:
        """Azzera i path completi dei CSV selezionati e resetta le label a 'non selezionato'."""
        stato["csv_paths"] = {"CSV1": None, "CSV2": None, "CSV3": None}

        def _w():
            # Mostra i nomi attesi (da intestazione) ma senza path selezionato
            app.lbl_csv1.config(text=f"{stato['csv_expected']['CSV1']}  — (non selezionato)", foreground="#666")
            app.lbl_csv2.config(text=f"{stato['csv_expected']['CSV2']}  — (non selezionato)", foreground="#666")
            app.lbl_csv3.config(text=f"{stato['csv_expected']['CSV3']}  — (non selezionato)", foreground="#666")
            # reset full_path interno delle PathLabel
            for lbl in (app.lbl_csv1, app.lbl_csv2, app.lbl_csv3):
                if hasattr(lbl, '_full_path'):
                    lbl._full_path = ""
        app.after(0, _w)

    def choose_excel() -> None:
        # cartella iniziale Excel: ultima usata (solo Excel) oppure default Temperature
        base = Path(getattr(settings, 'last_excel_dir', DEFAULT_EXCEL_DIR)).expanduser()
        initial_dir = str(base) if base.exists() else DEFAULT_EXCEL_DIR

        path = filedialog.askopenfilename(
            title="Scegli il file Excel di destinazione",
            initialdir=initial_dir,
            filetypes=[("Excel", "*.xlsx"), ("Tutti i file", "*.*")]
        )
        if not path:
            return

        xls_path = Path(path)

        # aggiorna percorsi persistenti
        settings.base_path = str(xls_path.parent)
        settings.last_excel_dir = str(xls_path.parent)
        save_settings(settings)

        stato["excel_path"] = str(xls_path)
        stato["log_file"] = xls_path.with_suffix(".log")
        log.info(f"Excel selezionato: {xls_path}")

        # aggiorna label excel
        app.lbl_excel.set_path(xls_path)

        # leggi intestazioni B1–D1 e aggiorna "nomi attesi"
        try:
            expected = read_expected_csv_names_from_excel(xls_path)
            stato["csv_expected"] = expected
            ui_log(f"📄 Excel selezionato: {log_path(xls_path)}")
            ui_log(f"🔎 Intestazioni CSV lette da B1–D1: {expected['CSV1']} | {expected['CSV2']} | {expected['CSV3']}")
        except Exception as e:
            # se fallisce, mantieni placeholder
            stato["csv_expected"] = {"CSV1": "CSV 1", "CSV2": "CSV 2", "CSV3": "CSV 3"}
            ui_log(f"⚠️ Non riesco a leggere le intestazioni B1–D1: {e}")

        # ✅ richiesta 2: appena scelto un nuovo Excel, azzera i path CSV precedenti
        reset_csv_state_keep_expected()

    def choose_csv(key: str, lbl_widget) -> None:
        if not stato["excel_path"]:
            messagebox.showinfo("Info", "Prima scegli il file Excel.")
            return

        # cartella iniziale CSV: ultima usata (solo CSV) oppure default Downloads
        base_csv = Path(getattr(settings, 'last_csv_dir', DEFAULT_CSV_DIR)).expanduser()
        initial_dir = str(base_csv) if base_csv.exists() else DEFAULT_CSV_DIR

        expected_name = stato["csv_expected"].get(key, key)

        path = filedialog.askopenfilename(
            title=f"Scegli il file CSV per {key} ({expected_name})",
            initialdir=str(initial_dir),
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return

        stato["csv_paths"][key] = path
        # aggiorna ultima cartella CSV (separata da Excel)
        try:
            settings.last_csv_dir = str(Path(path).parent)
            save_settings(settings)
        except Exception:
            pass
        # mostra basename selezionato, mantenendo il nome atteso
        lbl_widget.config(text=f"{expected_name}  — {safe_basename(path)}", foreground="#111")
        ui_log(f"✅ {key}: selezionato {log_path(path)}")

    def run() -> None:
        if not stato["excel_path"]:
            messagebox.showerror("Errore", "Seleziona prima il file Excel.")
            return

        selected = {k: v for k, v in stato["csv_paths"].items() if v}
        if not selected:
            messagebox.showerror("Errore", "Seleziona almeno un file CSV.")
            return

        # snapshot config per thread
        config = {
            "excel_path": stato["excel_path"],
            "csv_paths": dict(stato["csv_paths"]),
            "csv_expected": dict(stato["csv_expected"]),
        }

        ui_progress(0)
        ui_status("Avvio…")
        ui_log("🚀 Avvio caricamento…")

        th = threading.Thread(
            target=esegui_caricamento,
            args=(config, ui_log, ui_progress, ui_status),
            daemon=True
        )
        th.start()

    # -----------------------
    # Worker
    # -----------------------
    def esegui_caricamento(
        config: dict,
        log: Callable[[str], None],
        progress: Callable[[int], None],
        status: Callable[[str], None],
    ) -> None:
        # NB: il parametro 'log' e' la callback UI, non il logger IRC modulo-level.
        # Per questo qui non si usa il logger di sessione.
        try:
            status("Lettura CSV…")
            progress(5)

            dati_uniti: list[Tuple[str, dict]] = []
            csv_paths: Dict[str, Optional[str]] = config["csv_paths"]

            # ordina sempre CSV1–3
            for i, key in enumerate(("CSV1", "CSV2", "CSV3"), start=1):
                path = csv_paths.get(key)
                if not path:
                    continue
                nome_atteso = config["csv_expected"].get(key, key)
                log(f"📥 Lettura {key} ({nome_atteso}): {log_path(path)}")
                dati = lettura_csv_freeze.leggi_csv_4_0(path)
                dati_uniti.append((key, dati))
                progress(5 + i * 20)

            if not dati_uniti:
                raise ValueError("Nessun CSV valido da processare.")

            status("Scrittura su Excel…")
            progress(80)
            scrittura_excel_freeze.scrivi_excel_1_4(config["excel_path"], dati_uniti)

            progress(100)
            status("Completato ✅")
            log("✅ Scrittura completata.")
            log("✅ Caricamento completato con successo.")

        except Exception as e:
            status("Errore ❌")
            log(f"❌ Errore: {e}")

    # -----------------------
    # Bind methods expected by gui_freeze
    # -----------------------
    app.choose_excel = choose_excel
    app.choose_csv = choose_csv
    app.run = run

    # build GUI
    gui_freeze.crea_gui_1_4_2(app)

    # inizializza label CSV a placeholder coerente
    reset_csv_state_keep_expected()

    app.mainloop()


# ============================================================
#  ARGPARSE / ENTRY POINT
# ============================================================

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Caricamento Temperature — Freeze 1.4.2")
    p.add_argument(
        "--base_path",
        default=None,
        help="Cartella iniziale per scegliere Excel/CSV (persistente in config.json).",
    )
    return p.parse_args()


if __name__ == "__main__":
    import sys
    sys.dont_write_bytecode = True
    main()
